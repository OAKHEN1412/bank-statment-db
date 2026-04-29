"""
checker.py  —  Parse Excel invoice file and match against bank statement DB.

Column E (วันที่รับเงิน) format examples:
  '1/12/68 เวลา 12.13 น. ยอด 1,780.48 บาท\n#KBANK'
  '30/1/69 เวลา 03.08 น. ยอด 53,308.07 บาท\n#KTB ค่าธรรมเนียม 12 บาท'
  '25/2/69 เวลา 16.30 น. ยอด 151,882.24 บาท\n#KBANK'

Date uses Thai Buddhist Era short year: 68 → 2025 CE, 69 → 2026 CE
"""

import re
import io
from datetime import datetime, date as date_type
import openpyxl
from db import find_match

BANKS = ['BBL', 'KBANK', 'KTB', 'SCB', 'TTB']

# ──────────────────────────────────────────
#  Parse note from column E
# ──────────────────────────────────────────

def _be_to_ce(yy: int) -> int:
    """Convert Buddhist Era (short or full) to Christian Era."""
    if yy < 100:          # short: 68 → 2568 BE → 2025 CE
        return yy + 2500 - 543
    elif yy > 2400:       # full BE: 2568 → 2025 CE
        return yy - 543
    return yy             # already CE


def parse_note(note: str) -> dict | None:
    """
    Extract BANK, date (YYYY-MM-DD), time (HH:MM), amount from free-text note.
    Returns dict or None if cannot parse.
    """
    if not note or not isinstance(note, str):
        return None

    # ── Bank ──────────────────────────────
    bank_m = re.search(r'#\s*(BBL|KBANK|KTB|SCB|TTB)', note, re.IGNORECASE)
    bank = bank_m.group(1).upper() if bank_m else None

    # ── Date  DD/MM/YY or D/M/YY ─────────
    date_m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', note)
    parsed_date = None
    if date_m:
        d, m, y = int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3))
        try:
            parsed_date = datetime(
                _be_to_ce(y), m, d
            ).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # ── Time  HH.MM  ──────────────────────
    time_m = re.search(r'(\d{1,2})[.:](\d{2})\s*น\.', note)
    if not time_m:
        time_m = re.search(r'เวลา\s+(\d{1,2})[.:](\d{2})', note)
    parsed_time = None
    if time_m:
        h, mi = int(time_m.group(1)), int(time_m.group(2))
        parsed_time = f'{h:02d}:{mi:02d}'

    # ── Amount ────────────────────────────
    amount_m = re.search(r'ยอด\s*([\d,]+\.?\d*)\s*บาท', note)
    parsed_amount = None
    if amount_m:
        try:
            parsed_amount = float(amount_m.group(1).replace(',', ''))
        except ValueError:
            pass

    if not (bank and parsed_date and parsed_amount):
        return None

    return {
        'bank':   bank,
        'date':   parsed_date,             # YYYY-MM-DD
        'time':   parsed_time or '',       # HH:MM
        'amount': parsed_amount,
    }


# ──────────────────────────────────────────
#  Match one parsed note against DB
# ──────────────────────────────────────────

def match_row(parsed: dict) -> dict:
    """
    Try to find a matching DB transaction.
    Returns a result dict with verification flags.
    """
    if not parsed:
        return {
            'db': None,
            'bank_vf': False,
            'date_vf': False,
            'time_vf': False,
            'amount_vf': False,
            'matched': False,
        }

    db_row = find_match(
        parsed['bank'],
        parsed['date'],
        parsed['time'],
        parsed['amount'],
    )

    if not db_row:
        return {
            'db': None,
            'bank_vf': False,
            'date_vf': False,
            'time_vf': False,
            'amount_vf': False,
            'matched': False,
        }

    bank_vf   = (db_row['bank'] == parsed['bank'])
    date_vf   = (db_row['date'] == parsed['date'])
    time_vf   = (
        parsed['time'] != '' and
        db_row.get('time', '').startswith(parsed['time'])
    )
    db_amount = db_row.get('credit', 0) or db_row.get('debit', 0) or 0
    amount_vf = abs(db_amount - parsed['amount']) < 0.02

    return {
        'db':        db_row,
        'bank_vf':   bank_vf,
        'date_vf':   date_vf,
        'time_vf':   time_vf,
        'amount_vf': amount_vf,
        'matched':   bank_vf and date_vf and amount_vf,
    }


# ──────────────────────────────────────────
#  Process full Excel file
# ──────────────────────────────────────────

def _cell_value(cell):
    v = cell.value
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, date_type):
        return v.strftime('%d/%m/%Y')
    return v


def process_excel(file_stream) -> tuple[list[dict], dict]:
    """
    Read Excel file from a file-like stream.
    Returns (results_list, summary).
    Each item in results_list:
        seq, doc_no, buyer, date_invoice, note,
        parsed, match_result
    """
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
    ws = wb.active

    results = []
    total = matched = partial = no_parse = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Skip completely empty rows
        if all(c.value is None for c in row):
            continue

        seq          = row[0].value if len(row) > 0 else row_idx - 1
        date_invoice = _cell_value(row[1]) if len(row) > 1 else ''
        doc_no       = row[2].value if len(row) > 2 else ''
        buyer        = row[3].value if len(row) > 3 else ''
        note         = row[4].value if len(row) > 4 else ''

        parsed = parse_note(str(note) if note else '')
        if not parsed:
            no_parse += 1
            results.append({
                'seq': seq, 'doc_no': doc_no, 'buyer': buyer,
                'date_invoice': date_invoice, 'note': note,
                'parsed': None, 'match': None, 'status': 'no_parse',
            })
            total += 1
            continue

        mr = match_row(parsed)
        if mr['matched']:
            status = 'matched'
            matched += 1
        elif mr['db']:
            status = 'partial'
            partial += 1
        else:
            status = 'not_found'

        results.append({
            'seq': seq, 'doc_no': doc_no, 'buyer': buyer,
            'date_invoice': date_invoice, 'note': note,
            'parsed': parsed, 'match': mr, 'status': status,
        })
        total += 1

    not_found = total - matched - partial - no_parse
    summary = {
        'total':     total,
        'matched':   matched,
        'partial':   partial,
        'not_found': not_found,
        'no_parse':  no_parse,
        'rate':      round(matched / total * 100, 1) if total else 0,
    }
    wb.close()
    return results, summary


# ──────────────────────────────────────────
#  Export results to Excel
# ──────────────────────────────────────────

def export_excel(results: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Check Results'

    headers = [
        'ลำดับ', 'เลขที่เอกสาร', 'ชื่อผู้ซื้อ', 'วันที่ใบกำกับ', 'หมายเหตุรับเงิน',
        'BANK', 'Date', 'Time', 'Amount',
        'BANK_DB', 'Date_DB', 'Time_DB', 'Debit_DB', 'Credit_DB', 'Balance_DB',
        'BANK_VF', 'Date_VF', 'Time_VF', 'Amount_VF', 'Status',
    ]
    ws.append(headers)

    for r in results:
        p = r.get('parsed') or {}
        m = r.get('match') or {}
        db = m.get('db') or {}

        ws.append([
            r.get('seq', ''),
            r.get('doc_no', ''),
            r.get('buyer', ''),
            r.get('date_invoice', ''),
            r.get('note', ''),
            p.get('bank', ''),
            p.get('date', ''),
            p.get('time', ''),
            p.get('amount', ''),
            db.get('bank', ''),
            db.get('date', ''),
            db.get('time', ''),
            db.get('debit', ''),
            db.get('credit', ''),
            db.get('balance', ''),
            str(m.get('bank_vf', '')),
            str(m.get('date_vf', '')),
            str(m.get('time_vf', '')),
            str(m.get('amount_vf', '')),
            r.get('status', ''),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
